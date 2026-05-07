# 파일명: make_kau_json_xlsx.py
# 목적: 국내배출권(KAU).xlsx에서 대시보드용 ets_index_data.json 생성
#
# 사용 컬럼:
# - 날짜 / Date (Y/M/D)
# - 종가 / Closing
# - 대비 / Change
# - 등락률 / % Change
# - 거래량-종합 / Total Volume
#
# 출력 JSON 구조:
# [
#   {
#     "date": "2026-05-06",
#     "price": 12300,
#     "change": -100,
#     "change_rate": -0.81,
#     "volume": 1500
#   }
# ]
#
# 사용 방법:
# 1) 이 파일과 국내배출권(KAU).xlsx를 같은 폴더에 둡니다.
#    예: C:\dashboard\data\make_kau_json_xlsx.py
#        C:\dashboard\data\국내배출권(KAU).xlsx
# 2) 명령프롬프트에서 실행합니다.
#    python make_kau_json_xlsx.py
#
# 필요 패키지:
#   pip install openpyxl

from pathlib import Path
from datetime import datetime, date, timedelta
import json
import math
import re

from openpyxl import load_workbook


# =========================
# 1. 경로 설정
# =========================
BASE_DIR = Path(__file__).resolve().parent

INPUT_FILE = BASE_DIR / "국내배출권(KAU).xlsx"
OUTPUT_FILE = BASE_DIR / "ets_index_data.json"

# 최근 1년치만 저장
DAYS_TO_KEEP = 365


# =========================
# 2. 변환 함수
# =========================
def normalize_header(value):
    """헤더 비교를 쉽게 하기 위해 줄바꿈/공백/기호를 정리합니다."""
    if value is None:
        return ""

    text = str(value).strip().lower()
    text = text.replace("\n", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def find_column(headers, include_keywords, exclude_keywords=None):
    """
    헤더 목록에서 include_keywords를 모두 포함하고,
    exclude_keywords는 포함하지 않는 컬럼을 찾습니다.
    반환값은 0-based index입니다.
    """
    exclude_keywords = exclude_keywords or []

    for idx, header in enumerate(headers):
        if all(keyword.lower() in header for keyword in include_keywords):
            if not any(keyword.lower() in header for keyword in exclude_keywords):
                return idx

    return None


def parse_date(value):
    """엑셀 날짜 또는 문자열 날짜를 YYYY-MM-DD로 변환합니다."""
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip()
    if not text:
        return None

    text = text.replace(".", "/").replace("-", "/")
    text = re.sub(r"\s+", "", text)

    for fmt in ("%Y/%m/%d", "%y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass

    return None


def parse_number(value):
    """쉼표, %, +, ▲, ▼ 등이 포함된 값을 숫자로 변환합니다."""
    if value is None:
        return None

    if isinstance(value, float) and math.isnan(value):
        return None

    if isinstance(value, (int, float)):
        # 엑셀 숫자는 그대로 사용
        num = float(value)
    else:
        text = str(value).strip()

        if text in {"", "-", "NaN", "nan", "N/A", "n/a", "None"}:
            return None

        is_negative = False

        if "▼" in text or text.startswith("-") or text.startswith("−"):
            is_negative = True

        # 괄호 음수: (100)
        if text.startswith("(") and text.endswith(")"):
            is_negative = True
            text = text[1:-1]

        text = (
            text.replace(",", "")
                .replace("%", "")
                .replace("▲", "")
                .replace("▼", "")
                .replace("+", "")
                .replace("−", "-")
                .strip()
        )

        try:
            num = float(text)
        except ValueError:
            return None

        if is_negative and num > 0:
            num = -num

    # 123.0은 123으로 저장
    if float(num).is_integer():
        return int(num)

    return round(num, 4)


# =========================
# 3. 메인 로직
# =========================
def main():
    if not INPUT_FILE.exists():
        raise FileNotFoundError(
            f"입력 파일을 찾을 수 없습니다: {INPUT_FILE}\n"
            "make_kau_json_xlsx.py와 국내배출권(KAU).xlsx를 같은 폴더에 두었는지 확인해 주세요."
        )

    wb = load_workbook(INPUT_FILE, data_only=True)
    ws = wb.active

    # 이 파일은 1행이 헤더입니다.
    header_row = 1
    headers = [
        normalize_header(ws.cell(header_row, col).value)
        for col in range(1, ws.max_column + 1)
    ]

    # 실제 변환 대상 컬럼 찾기
    col_date = find_column(headers, ["날짜"])
    if col_date is None:
        col_date = find_column(headers, ["date"])

    col_price = find_column(headers, ["종가"])
    if col_price is None:
        col_price = find_column(headers, ["closing"])

    # '대비 Change'는 '% Change'와 구분해야 하므로 '등락률', '%' 제외
    col_change = find_column(headers, ["대비"])
    if col_change is None:
        col_change = find_column(headers, ["change"], exclude_keywords=["%", "등락률"])

    col_change_rate = find_column(headers, ["등락률"])
    if col_change_rate is None:
        col_change_rate = find_column(headers, ["% change"])

    # 거래량은 '거래량-종합 / Total Volume'을 우선 사용
    col_volume = find_column(headers, ["거래량-종합"])
    if col_volume is None:
        col_volume = find_column(headers, ["total volume"])
    if col_volume is None:
        # 최후 대안: 거래량이 들어간 컬럼 중 가장 마지막 컬럼 사용
        volume_candidates = [
            idx for idx, header in enumerate(headers)
            if "거래량" in header or "volume" in header
        ]
        col_volume = volume_candidates[-1] if volume_candidates else None

    required = {
        "날짜": col_date,
        "종가": col_price,
        "대비": col_change,
        "등락률": col_change_rate,
        "거래량-종합": col_volume,
    }

    missing = [name for name, idx in required.items() if idx is None]
    if missing:
        raise RuntimeError(
            "필수 컬럼을 찾지 못했습니다: " + ", ".join(missing) + "\n"
            "현재 인식한 헤더:\n- " + "\n- ".join(headers)
        )

    rows = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]

        date_value = parse_date(values[col_date])
        if not date_value:
            continue

        item = {
            "date": date_value,
            "price": parse_number(values[col_price]),
            "change": parse_number(values[col_change]),
            "change_rate": parse_number(values[col_change_rate]),
            "volume": parse_number(values[col_volume]),
        }

        # 가격이 없으면 대시보드 표시 대상에서 제외
        if item["price"] is None:
            continue

        rows.append(item)

    if not rows:
        raise RuntimeError("변환 가능한 데이터가 없습니다.")

    # 최신일자 우선 정렬
    rows.sort(key=lambda x: x["date"], reverse=True)

    # 최신일자 기준 최근 1년치만 추출
    latest_date = datetime.strptime(rows[0]["date"], "%Y-%m-%d").date()
    start_date = latest_date - timedelta(days=DAYS_TO_KEEP)

    rows_1y = [
        row for row in rows
        if datetime.strptime(row["date"], "%Y-%m-%d").date() >= start_date
    ]

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(rows_1y, f, ensure_ascii=False, indent=2)

    print(f"저장 완료: {OUTPUT_FILE}")
    print(f"기간: {rows_1y[-1]['date']} ~ {rows_1y[0]['date']}")
    print(f"건수: {len(rows_1y):,}건")

    print("\n미리보기:")
    for item in rows_1y[:5]:
        print(item)


if __name__ == "__main__":
    main()
