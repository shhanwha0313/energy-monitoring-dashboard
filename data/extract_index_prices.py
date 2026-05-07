import json
from datetime import datetime
from pathlib import Path
import openpyxl

BASE_DIR = Path(__file__).resolve().parent
INPUT_XLSX = BASE_DIR / 'issue monitering.xlsx'
OUTPUT_JSON = BASE_DIR / 'latest_prices.json'


def safe_pct_change(current, previous):
    if current is None or previous in (None, 0):
        return None
    return (current - previous) / previous * 100


def fmt_date(value):
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    return str(value)


def is_real_date(value):
    return isinstance(value, datetime)


def main():
    wb = openpyxl.load_workbook(INPUT_XLSX, data_only=True)
    ws = wb['Index']

    rows = []
    for row in range(4, ws.max_row + 1):
        date_value = ws[f'A{row}'].value
        smp = ws[f'B{row}'].value
        ets = ws[f'D{row}'].value
        rec = ws[f'F{row}'].value

        if not is_real_date(date_value):
            continue
        if smp is None and ets is None and rec is None:
            continue

        rows.append({
            'date': fmt_date(date_value),
            'smp': smp,
            'ets': ets,
            'rec': rec,
        })

    if not rows:
        raise ValueError('Index 시트에서 유효한 가격 데이터를 찾지 못했습니다.')

    latest = rows[-1]
    prev = rows[-2] if len(rows) >= 2 else {'smp': None, 'ets': None, 'rec': None}

    payload = {
        'dates': [r['date'] for r in rows],
        'smp': [r['smp'] for r in rows],
        'ets': [r['ets'] for r in rows],
        'rec': [r['rec'] for r in rows],
        'summary': {
            'latestDate': latest['date'],
            'smp': {
                'value': latest['smp'],
                'change': None if latest['smp'] is None or prev['smp'] is None else latest['smp'] - prev['smp'],
                'changePct': safe_pct_change(latest['smp'], prev['smp']),
                'unit': '원/kWh',
            },
            'ets': {
                'value': latest['ets'],
                'change': None if latest['ets'] is None or prev['ets'] is None else latest['ets'] - prev['ets'],
                'changePct': safe_pct_change(latest['ets'], prev['ets']),
                'unit': '원/tCO2',
            },
            'rec': {
                'value': latest['rec'],
                'change': None if latest['rec'] is None or prev['rec'] is None else latest['rec'] - prev['rec'],
                'changePct': safe_pct_change(latest['rec'], prev['rec']),
                'unit': '원/REC',
            },
        },
    }

    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f'생성 완료: {OUTPUT_JSON}')


if __name__ == '__main__':
    main()
